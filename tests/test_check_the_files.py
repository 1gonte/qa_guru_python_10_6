import csv
import pypdf

from openpyxl import load_workbook

from tests.conftest import PROJECT_RESOURCES_PATH


def test_file_xlsx_after_archivate():
    lw = load_workbook(PROJECT_RESOURCES_PATH + '/' + 'excel_file.xlsx')
    sheet = lw.active
    assert sheet.cell(1, 1).value == 'First Name'
    assert sheet.cell(1, 2).value == 'Last Name'
    assert sheet.cell(2, 1).value == 'Georgii'
    assert sheet.cell(2, 2).value == 'Sergeev'


def test_file_csv_after_archivate():
    with open(PROJECT_RESOURCES_PATH + '/' + 'csv_file.csv', 'r') as f:
        reader = csv.reader(f)
        assert next(reader)[0] == '5GD1LBUQ15JTV7QN'


def test_file_pdf_after_archivate():
    reader = pypdf.PdfReader(PROJECT_RESOURCES_PATH + '/' + 'pdf_file.pdf')
    assert "Copeland" in reader.pages[0].extract_text()
